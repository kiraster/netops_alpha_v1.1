'''
全局设置和共用函数
'''
import os
from datetime import datetime
from functools import wraps
import ipaddress
from openpyxl import load_workbook
from multiprocessing.pool import ThreadPool
from netmiko import ConnectHandler, NetMikoAuthenticationException, NetMikoTimeoutException, ssh_exception


# 项目根目录
BASE_PATH = os.path.dirname((__file__))

# 文件输出目录
EXPORT_PATH = os.path.join(BASE_PATH, 'EXPORT')

# 定义目录名称为当天日期（格式：20220609）
dir_name = datetime.now().strftime("%Y%m%d")

# 目录创建
new_path = os.path.join(EXPORT_PATH, dir_name)
if not os.path.isdir(new_path):
    os.makedirs(new_path)

backup_path = os.path.normpath(
    os.path.join(EXPORT_PATH, dir_name, 'config_backup'))
config_path = os.path.normpath(
    os.path.join(EXPORT_PATH, dir_name, 'config_add'))
# generate_table = os.path.normpath(
#     os.path.join(EXPORT_PATH, dir_name, 'generate_table'))

if not os.path.isdir(backup_path):
    os.makedirs(backup_path)
if not os.path.isdir(config_path):
    os.makedirs(config_path)
# if not os.path.isdir(generate_table):
#     os.makedirs(generate_table)

# 表格数据文件路径
device_file = "dev_data.xlsx"

# ThreadPool 设定异步进程数为66
t_pool = ThreadPool(66)

# 定义列表保存执行成功和失败的主机IP
sucessful_list = []
failed_list = []


# 手动输入登陆信息
def login_start():
    # login module，校验由netmiko完成
    # login_user = input('Login:')
    # login_pwd = getpass.getpass('Passwd:')
    # # 明文写死在代码，不安全的方式
    # login_user = 'admin'
    # login_pwd = 'xxx'
    # return login_user, login_pwd
    pass


# 加载excel文件
def load_excel():
    try:
        wb = load_workbook(device_file)
        return wb
    except FileNotFoundError:
        print("{} excel文件不存在".format(device_file))
    except Exception:
        print("载入读取{} excel文件失败".format(device_file))


# 获取设备数据信息
def get_device_info(task_name):
    try:
        # by openpyxl
        # user, pwd = login_start()
        wb = load_excel()
        ws1 = wb[wb.sheetnames[0]]
        # 选定单元格数据区域
        for row in ws1.iter_rows(min_row=2, max_col=9):
            # 判断IP所在的列不为空值，则进行如下代码
            if row[2].value:
                if str(row[1].value).strip() == '#':
                    continue
                info_dict = {
                    'ip':
                    row[2].value,
                    'username':
                    row[5].value,
                    'password':
                    row[6].value,
                    'protocol':
                    row[3].value,
                    'port':
                    row[4].value,
                    'secret':
                    row[7].value,
                    'device_type':
                    row[8].value,
                    'cmd_list':
                    get_cmd_info(task_name, wb[row[8].value.strip().lower()]),
                }
                yield info_dict
            else:
                break
    except Exception as e:
        print("get_device_info failed: {}".format(e))
    finally:
        wb.close()


# 获取命令信息
def get_cmd_info(task_name, sheet_name):
    cmd_list = []
    try:
        # by openpyxl
        for row in sheet_name.iter_rows(min_row=2, max_col=3):
            # 若单元格使用“＃”进行注释或命令为空值，则跳过该行
            if str(row[0].value).strip() != '#' and row[1].value and task_name == 'backup_config.py':
                cmd_list.append(row[1].value.strip())
            elif str(row[0].value).strip() != '#' and row[2].value and task_name == 'add_config.py':
                cmd_list.append(row[2].value.strip())
        return cmd_list
    except Exception as e:
        print("get_cmd_info Error: ", e)


# 获取自定义设备数据信息
def get_undifined_device_info():
    try:
        # by openpyxl
        wb = load_excel()
        ws1 = wb[wb.sheetnames[5]]
        row_number = 2
        # 选定单元格数据区域
        for row in ws1.iter_rows(min_row=2, max_col=9):
            # 获取执行命令
            undifined_cmd_info = []
            for cols in ws1.iter_cols(min_col=10,
                                        min_row=row_number,
                                        max_row=row_number,
                                        values_only=True):
                for col in cols:
                    if col is None:
                        continue
                    undifined_cmd_info.append(col)
            row_number += 1

            # 判断IP所在的列不为空值，执行如下代码
            if row[2].value:
                if str(row[1].value).strip() == '#':
                    continue
                info_dict = {
                    'ip':
                    row[2].value,
                    'username':
                    row[5].value,
                    'password':
                    row[6].value,
                    'protocol':
                    row[3].value,
                    'port':
                    row[4].value,
                    'secret':
                    row[7].value,
                    'device_type':
                    row[8].value,
                    'cmd_list': undifined_cmd_info,
                }
                yield info_dict
            else:
                break
    except Exception as e:
        print("get_device_info failed: {}".format(e))
    finally:
        wb.close()


# netmiko 连接处理
def connect_handler(host):
    try:
        connect = ''
        # 判断使用SSH协议
        # 将“protocol”列单元格的内容大写转小写，去除前后空格对比是否为 ssh
        if host['protocol'].lower().strip() == 'ssh':
            #  判断“port”列单元格，若单元格填写的内容不是22或空，则定义为22
            host['port'] = host['port'] if (host['port'] not in [22, None]) else 22
            # 剔除多余connectHandler 不需要的参数，protocol、secret、cmd_list，华为华为设备无特权密码
            host.pop('protocol'), host.pop('cmd_list')

            if 'huawei' in host['device_type']:
                host.pop('secret')
                connect = ConnectHandler(**host, conn_timeout=10)
            elif 'hp_comware' in host['device_type']:
                host.pop('secret')
                connect = ConnectHandler(**host)
            else:
                connect = ConnectHandler(**host, conn_timeout=10)
        # 判断使用Telnet协议
        elif host['protocol'].lower().strip() == 'telnet':
            # 判断“port”列单元格，若单元格填写的内容不是23或空，则定义为23
            host['port'] = host['port'] if (host['port'] not in [23, None]) else 23
            # 剔除多余connectHandler 不需要的参数，protocol、secret、cmd_list，华三华为设备无特权密码
            host.pop('protocol'), host.pop('secret'), host.pop('cmd_list')
            # netmiko 支持telnet协议，设备类型格式为：hp_comware_telnet
            host['device_type'] = host['device_type'] + '_telnet'
            # fast_cli=false,待测试参数
            connect = ConnectHandler(**host, fast_cli=False)
        # 不在以上两种协议内的连接
        else:
            res = "暂不支持IP地址为{}_的设备使用{}协议登陆".format(host['ip'], host['protocol'])
            raise ValueError(res)
        return True, connect

    except NetMikoTimeoutException:
        res = "{} Can not connect to Device!".format(host['ip'])
        print(res)
        return False, res
    except NetMikoAuthenticationException:
        res = "{} username/password wrong!".format(host['ip'])
        print(res)
        return False, res
    except ssh_exception:
        res = "{} SSH parameter problem!".format(host['ip'])
        print(res)
        return False, res
    except Exception as e:
        print("{} Failed: {}".format(host['ip'], e))
        return False, res


# 记录程序执行时间装饰器
def timer(func):

    @wraps(func)
    def wrapper(*args, **kwargs):
        start_time = datetime.now()
        func(*args, **kwargs)
        end_time = datetime.now()
        # print('\n' + '-' * 42)
        print('执行完毕，共耗时 {:0.2f} 秒.'.format((end_time - start_time).total_seconds()))
        print('-' * 42)
        # return res

    return wrapper


# 记录运行结果的装饰器
def result_count(func):

    @wraps(func)
    def wrapper(*args, **kwargs):
       
        task_name, sucessful_list, failed_list = func(*args, **kwargs)
        result_count = ('设备总数 {} 台，成功 {} 台，失败 {} 台.'.format(
            len(sucessful_list) + len(failed_list), len(sucessful_list), len(failed_list)))
        print('\n' + '-' * 42)
        print(result_count)
        result_path = os.path.normpath(os.path.join(EXPORT_PATH, dir_name, f'result_{dir_name}.log'))
        print(f'\n运行结果保存路径: \"{result_path}\"\n')

        return task_name, sucessful_list, failed_list

    return wrapper


# 写入文件
def write_to_file(task_name, output_filename, output):
    # 写入结果到文件
    if task_name == 'backup_config.py':
        with open(os.path.join(backup_path, output_filename), 'a', encoding="GB18030") as f:
            f.write(output)
    elif task_name == 'add_config.py':
        with open(os.path.join(config_path, output_filename), 'a', encoding="GB18030") as f:
            f.write(output)
    elif task_name == 'undifined.py':
        with open(os.path.join(config_path, output_filename), 'a', encoding="GB18030") as f:
            f.write(output)
    else:
        pass


# 保存运行结果记录的装饰器
def result_write(func):

    @wraps(func)
    def wrapper(*args, **kwargs):

        task_name, sucessful_list, failed_list = func(*args, **kwargs)
        result_count = ('设备总数 {} 台，成功 {} 台，失败 {} 台.'.format(
            len(sucessful_list) + len(failed_list), len(sucessful_list), len(failed_list)))

        # time_str = datetime.now()
        time_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

        result_path = os.path.normpath(
            os.path.join(EXPORT_PATH, dir_name, f'result_{dir_name}.log'))

        with open(result_path, 'a', encoding="utf-8") as f:
            log_title = task_name.center(100, '=') + '\n' + time_str.center(100, '=') + '\n'
            f.write(log_title)
            f.write(result_count + '\n')
            f.write('\n执行成功设备列表：\n')
            for i in sucessful_list:
                f.write(i)
                f.write('\n')

            f.write('\nNG设备列表：\n')
            for i in failed_list:
                f.write(i)
                f.write('\n')
            f.write('\n')

        return task_name, sucessful_list, failed_list

    return wrapper


# 判断是否是正确格式的IP地址，IP地址网络，IP地址范围
def is_valid_ipv4_input(ipv4_str):

    try:
        # 尝试将输入解析为 IPv4Address
        ipaddress.IPv4Address(ipv4_str)
        return True
    except ValueError:
        return False
        # try:
        #     # 将输入拆分为两个 IP 地址
        #     start, end = ipv4_str.split('-')
        #     # 尝试将输入解析为 IPv4Address
        #     ipaddress.IPv4Address(start.strip())
        #     ipaddress.IPv4Address(end.strip())
        #     return True
        # except ValueError as e:
        #     # 尝试将输入解析为 IPv4Network
        #     try:
        #         network = ipaddress.IPv4Network(ipv4_str)
        #         if network.hostmask != '0.0.0.0':
        #             return True
        #         else:
        #             return False
        #     except ValueError as e:
        #         # print(e)
        #         return False